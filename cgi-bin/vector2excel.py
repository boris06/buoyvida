import cgi
import cgitb
import sys
import numpy as np

from buoydef import *
from buoyvida import *

from openpyxl import Workbook
from openpyxl.cell import get_column_letter

from StringIO import StringIO
    
cgitb.enable()

fs = cgi.FieldStorage()

##print "Content-Type: text/html"
##print ""

startDateTime = fs["startDateTime"].value
endDateTime = fs["endDateTime"].value
exec('cells =' + fs["cells"].value)
exec('heights =' + fs["heights"].value)

##startDateTime = '2015-09-03T00:00:00'
##endDateTime = '2015-09-05T23:30:00'
##

# make period list
period_list = make_period_list(startDateTime,endDateTime)

# get wind and waves data
fields=['wind.vmspd', 'wind.vmdir', 'awac_waves.wave_height', 'awac_waves.mean_direction']
tables=['wind', 'awac_waves']
# table
fieldDesc = [fieldDict[x] for x in fields]

(nseries,series)=get_buoy_data(fields,tables,period_list)

# get sea currents data
(current_E,current_N) = get_buoy_currents(period_list,cells)
uvec = np.vstack(current_E[0])
uvec = uvec.T
for i in range(1,len(current_E)):
    uvec = np.vstack((uvec,current_E[i].T))                     
vvec = np.vstack(current_N[0])
vvec = vvec.T
for i in range(1,len(current_N)):
    vvec = np.vstack((vvec,current_N[i].T))
    
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws.cell(row = 1, column = 1).value = "Date and time"
for i in range(nseries):
    ws.cell(row = 1, column = i+2).value = fieldDesc[i]
fieldDescCurr = ['Current%s (%s m)' % (direction,height) for height in heights for direction in ['E','N']]
for i in range(len(fieldDescCurr)):
    ws.cell(row = 1, column = i+2+nseries).value = fieldDescCurr[i]

for row in range(2,len(period_list)+2):
    ws.cell(row = row, column = 1).value = period_list[row-2]
    ws.cell(row = row, column = 1).number_format = 'd.m.yyyy h:mm'
    for col in range(nseries):
        if (np.isnan(series[row-2,col])):
            ws.cell(row = row, column = col+2).value = None
        else:
            ws.cell(row = row, column = col+2).value = series[row-2,col]
            ws.cell(row = row, column = col+2).number_format = '0.00'
    count = 0
    for col in range(len(cells)):
        if (np.isnan(uvec[col,row-2])):
            ws.cell(row = row, column = 2+nseries+count).value = None
            ws.cell(row = row, column = 2+nseries+count+1).value = None
        else:
            ws.cell(row = row, column = 2+nseries+count).value = uvec[col,row-2]
            ws.cell(row = row, column = 2+nseries+count+1).value = vvec[col,row-2]
        count = count + 2
        
for i in range(1+nseries+len(cells)*2):
   ws.column_dimensions[get_column_letter(i+1)].width = 18 

start_date = period_list[0].strftime('%Y-%m-%d %H:%M:%S')    
end_date = period_list[-1].strftime('%Y-%m-%d %H:%M:%S')    

wbname = 'wind_waves_currents_' + ("%s_%s.xlsx" % (start_date,end_date))
wbname = wbname.replace("-","")
wbname = wbname.replace(":","")
wbname = wbname.replace(" ","_")    

print 'Content-Type: application/octet-stream'
print 'Content-Disposition: attachment; filename="%s"' % wbname
print

output = StringIO()
wb.save(output)
sys.stdout.write(output.getvalue())
sys.stdout.flush()
