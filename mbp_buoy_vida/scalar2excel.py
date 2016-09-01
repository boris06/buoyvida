#!/usr/bin/env python 
# -*- coding: UTF-8 -*- 

__DBG = False
if __DBG == True:
    import cgitb
    cgitb.enable()            # Enable detailed and formated exception stacktrace logging

import numpy as np

from buoydef import *
from buoyvida import *

from openpyxl import Workbook

from StringIO import StringIO

from mbp_buoy_vida import config as DbConfig
dbConfig = DbConfig.DbConfig()


def scalar2excel(selectPlot=None, startDateTime=None, endDateTime=None):
    global dbConfig

    ##selectPlot = "Wind Speed and Gusts"
    ##startDateTime = '2015-09-03T00:00:00'
    ##endDateTime = '2015-09-05T23:30:00'
    ##

    # make period list
    period_list = make_period_list(startDateTime,endDateTime)

    # get buoy database parameters
    (fields,tables,fieldDesc,yLab,axis,colorStyle) = get_buoyparam(selectPlot)

    # get buoy scalar data
    (nseries,series) = get_buoy_data(dbConfig, fields, tables, period_list)

    wb = Workbook()

    #ws = wb.active       # NOTE: not available in ver 1.7
    ws = wb.get_active_sheet()

    # Data can be assigned directly to cells
    ws.cell(row = 1, column = 1).value = "Date and time"
    for i in range(nseries):
        ws.cell(row = 1, column = i+2).value = fieldDesc[i]

    for row in range(2,len(period_list)+2):
        ws.cell(row = row, column = 1).value = period_list[row-2]
        #ws.cell(row = row, column = 1).number_format = 'd.m.yyyy h:mm'     # NOTE: not available in ver 1.7
        for col in range(nseries):
            if (np.isnan(series[row-2,col])):
                ws.cell(row = row, column = col+2).value = None
            else:
                ws.cell(row = row, column = col+2).value = series[row-2,col]
                # ws.cell(row = row, column = col+2).number_format = '0.00' # NOTE: not available in ver 1.7

    # ???
    # colletters = map(chr, range(65, 65+nseries+1))
    # for colletter in colletters:
    #     ws.column_dimensions[colletter].width = 18

    start_date = period_list[0].strftime('%Y-%m-%d %H:%M:%S')    
    end_date = period_list[-1].strftime('%Y-%m-%d %H:%M:%S')    

    fieldDesc1 = [x.replace(" ","") for x in fieldDesc]
    wbname = "_".join(fieldDesc1) + "_" + ("%s_%s.xlsx" % (start_date,end_date))
    wbname = wbname.replace("-","")
    wbname = wbname.replace(":","")
    wbname = wbname.replace(" ","_")    

    output = StringIO()
    wb.save(output)
    return [wbname, output.getvalue()]
