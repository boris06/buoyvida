#!/usr/bin/env python 
# -*- coding: UTF-8 -*- 

__DBG = False
if __DBG == True:
    import cgitb
    cgitb.enable()            # Enable detailed and formated exception stacktrace logging

import numpy as np

from buoydef import *
from buoyvida import *

from openpyxl import Workbook
from openpyxl.cell import get_column_letter

import string
from StringIO import StringIO

from mbp_buoy_vida import config as DbConfig
dbConfig = DbConfig.DbConfig()


def vector2excel(selectHeights=None, startDateTime=None, endDateTime=None):
    global dbConfig
    all_heights = ['%d' % height for height in range(2,23)]
    def_heights = ['2', '5', '10', '15', '20']
    if selectHeights is None  or  startDateTime is None  or  endDateTime is None:
        endDateTime = datetime.today().strftime('%Y-%m-%dT%H:%M:%S')
        startDateTime = (datetime.today() - timedelta(days=2)).strftime('%Y-%m-%dT%H:%M:%S')
        heights = def_heights
    else:
        try:
            heights = selectHeights.split(',')
        except:
            heights = selectHeights

    start_date = startDateTime.replace('T', ' ')
    end_date = endDateTime.replace('T', ' ')

    cells = map(str, (np.asarray(map(int, heights)) - 2))

    # make period list
    period_list = make_period_list(startDateTime, endDateTime)

    # get wind and waves data
    fields=['wind.vmspd', 'wind.vmdir', 'awac_waves.wave_height', 'awac_waves.mean_direction']
    tables=['wind', 'awac_waves']
    # table
    # ??? fieldDesc = [fieldDict[x] for x in fields]
    fieldDesc = [fieldDict[x] for x in fields]
    # fieldDesc = string.join(fields)

    (nseries,series)=get_buoy_data(dbConfig, fields, tables, period_list)

    # get sea currents data
    (current_E,current_N) = get_buoy_currents(dbConfig, period_list, cells)
    uvec = np.vstack(current_E[0])
    uvec = uvec.T
    for i in range(1,len(current_E)):
        uvec = np.vstack((uvec,current_E[i].T))                     
    vvec = np.vstack(current_N[0])
    vvec = vvec.T
    for i in range(1,len(current_N)):
        vvec = np.vstack((vvec,current_N[i].T))
        
    wb = Workbook()

    #ws = wb.active       # NOTE: not available in ver 1.7
    ws = wb.get_active_sheet()

    # Data can be assigned directly to cells
    ws.cell(row = 1, column = 1).value = "Date and time"
    for i in range(nseries):
        ws.cell(row = 1, column = i+2).value = fieldDesc[i]
    fieldDescCurr = ['Current%s (%s m)' % (direction,height) for height in heights for direction in ['E','N']]
    for i in range(len(fieldDescCurr)):
        ws.cell(row = 1, column = i+2+nseries).value = fieldDescCurr[i]

    for row in range(2,len(period_list)+2):
        ws.cell(row = row, column = 1).value = period_list[row-2]
        #ws.cell(row = row, column = 1).number_format = 'd.m.yyyy h:mm'     # NOTE: not available in ver 1.7
        for col in range(nseries):
            if (np.isnan(series[row-2,col])):
                ws.cell(row = row, column = col+2).value = None
            else:
                ws.cell(row = row, column = col+2).value = series[row-2,col]
                # ws.cell(row = row, column = col+2).number_format = '0.00' # NOTE: not available in ver 1.7
        count = 0
        for col in range(len(cells)):
            if (np.isnan(uvec[col,row-2])):
                ws.cell(row = row, column = 2+nseries+count).value = None
                ws.cell(row = row, column = 2+nseries+count+1).value = None
            else:
                ws.cell(row = row, column = 2+nseries+count).value = uvec[col,row-2]
                ws.cell(row = row, column = 2+nseries+count+1).value = vvec[col,row-2]
            count = count + 2

    # ???
    # for i in range(1+nseries+len(cells)*2):
    #    ws.column_dimensions[get_column_letter(i+1)].width = 18 

    start_date = period_list[0].strftime('%Y-%m-%d %H:%M:%S')    
    end_date = period_list[-1].strftime('%Y-%m-%d %H:%M:%S')    

    wbname = 'wind_waves_currents_' + ("%s_%s.xlsx" % (start_date,end_date))
    wbname = wbname.replace("-","")
    wbname = wbname.replace(":","")
    wbname = wbname.replace(" ","_")    

    output = StringIO()
    wb.save(output)
    return [wbname, output.getvalue()]
